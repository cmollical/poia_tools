import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models import ClassifiedIntent, EpicMatch

class ExcelExportService:
    def __init__(self, export_directory: str = "exports"):
        self.export_directory = export_directory
        self.excel_filename = "project_stories.xlsx"
        self.excel_path = os.path.join(export_directory, self.excel_filename)
        
        # Create exports directory if it doesn't exist
        os.makedirs(export_directory, exist_ok=True)
        
        # Initialize Excel file if it doesn't exist
        if not os.path.exists(self.excel_path):
            self._create_initial_excel_file()

    def _create_initial_excel_file(self):
        """Create initial Excel file with headers."""
        # Define columns for our export
        columns = [
            "Timestamp",
            "Story ID",
            "Type",
            "Priority", 
            "Summary",
            "Description",
            "Acceptance Criteria",
            "Epic Keywords",
            "Matched Epic ID",
            "Matched Epic Name",
            "Epic Match Confidence",
            "AI Confidence",
            "Processing Time (ms)",
            "Cleaned Transcript",
            "Original Transcript"
        ]
        
        # Create DataFrame with headers
        df = pd.DataFrame(columns=columns)
        
        # Create workbook and write data
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Stories"
        
        # Write headers
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        column_widths = [20, 12, 10, 10, 40, 60, 50, 30, 15, 25, 15, 12, 15, 60, 60]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        wb.save(self.excel_path)

    def export_story(
        self,
        classified_intent: ClassifiedIntent,
        epic_match: Optional[EpicMatch],
        cleaned_transcript: str,
        original_transcript: str,
        processing_time_ms: int
    ) -> Dict[str, Any]:
        """
        Export a single story to Excel file.
        Each story becomes a new row in the spreadsheet.
        """
        try:
            # Generate unique story ID based on timestamp
            timestamp = datetime.now()
            story_id = f"STORY-{timestamp.strftime('%Y%m%d-%H%M%S')}"
            
            # Prepare story data
            story_data = {
                "Timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                "Story ID": story_id,
                "Type": classified_intent.type.value.upper(),
                "Priority": classified_intent.priority.value.upper(),
                "Summary": classified_intent.summary,
                "Description": classified_intent.description,
                "Acceptance Criteria": " | ".join(classified_intent.acceptance_criteria) if classified_intent.acceptance_criteria else "",
                "Epic Keywords": ", ".join(classified_intent.epic_keywords) if classified_intent.epic_keywords else "",
                "Matched Epic ID": epic_match.epic_id if epic_match and epic_match.epic_id else "",
                "Matched Epic Name": epic_match.epic_name if epic_match and epic_match.epic_name else "",
                "Epic Match Confidence": f"{epic_match.match_confidence:.2%}" if epic_match else "0%",
                "AI Confidence": f"{classified_intent.confidence:.2%}",
                "Processing Time (ms)": processing_time_ms,
                "Cleaned Transcript": cleaned_transcript,
                "Original Transcript": original_transcript
            }
            
            # Load existing workbook
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Find the next empty row
            next_row = ws.max_row + 1
            
            # Write data to the new row
            for col_num, (column, value) in enumerate(story_data.items(), 1):
                cell = ws.cell(row=next_row, column=col_num)
                cell.value = value
                
                # Apply conditional formatting based on type and priority
                if column == "Type":
                    if value == "BUG":
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif value == "STORY":
                        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    elif value == "EPIC":
                        cell.fill = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")
                    elif value == "TASK":
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
                elif column == "Priority":
                    if value == "CRITICAL":
                        cell.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif value == "HIGH":
                        cell.fill = PatternFill(start_color="FF9966", end_color="FF9966", fill_type="solid")
                    elif value == "MEDIUM":
                        cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
                
                # Set text wrapping for longer text fields
                if column in ["Description", "Acceptance Criteria", "Cleaned Transcript", "Original Transcript"]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")
            
            # Set row height for better readability
            ws.row_dimensions[next_row].height = 60
            
            # Save the workbook
            wb.save(self.excel_path)
            
            return {
                "success": True,
                "story_id": story_id,
                "excel_path": self.excel_path,
                "row_number": next_row,
                "message": f"Story {story_id} exported to Excel successfully"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to export story to Excel: {str(e)}"
            }

    def get_export_summary(self) -> Dict[str, Any]:
        """Get summary information about exported stories."""
        try:
            if not os.path.exists(self.excel_path):
                return {
                    "total_stories": 0,
                    "excel_exists": False,
                    "excel_path": self.excel_path
                }
            
            # Load workbook and count rows (minus header)
            wb = load_workbook(self.excel_path)
            ws = wb.active
            total_stories = ws.max_row - 1 if ws.max_row > 1 else 0
            
            # Get file modification time
            file_stats = os.stat(self.excel_path)
            last_modified = datetime.fromtimestamp(file_stats.st_mtime)
            
            return {
                "total_stories": total_stories,
                "excel_exists": True,
                "excel_path": self.excel_path,
                "last_modified": last_modified.strftime("%Y-%m-%d %H:%M:%S"),
                "file_size_kb": round(file_stats.st_size / 1024, 2)
            }
            
        except Exception as e:
            return {
                "total_stories": 0,
                "excel_exists": False,
                "excel_path": self.excel_path,
                "error": str(e)
            }

    def get_excel_file_path(self) -> str:
        """Get the full path to the Excel file."""
        return os.path.abspath(self.excel_path)
